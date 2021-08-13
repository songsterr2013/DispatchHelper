from os import listdir
import os.path
import sys
from json import load
from operator import itemgetter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from xlrd import open_workbook
from mysql.connector import connect, Error


class ReadExcel:

    def __init__(self, db_file_1, db_file_2):

        self.root_path = ''

        if getattr(sys, 'frozen', False):  # 如果為exe
            self.root_path = os.path.dirname(sys.executable)
        elif __file__:
            self.root_path = os.path.dirname(__file__)

        self.previous_root_path = os.path.dirname(self.root_path)  # 取得檔案所在的前一路徑

        if not os.path.isdir(os.path.join(self.previous_root_path, "1_ERP產品_excel")):  # 不存在就創folder
            os.mkdir(os.path.join(self.previous_root_path, "1_ERP產品_excel"))
        if not os.path.isdir(os.path.join(self.previous_root_path, "2_炸製令結果_excel")):  # 不存在就創folder
            os.mkdir(os.path.join(self.previous_root_path, "2_炸製令結果_excel"))
        if not os.path.isdir(os.path.join(self.previous_root_path, "3_派工單_excel")):  # 不存在就創folder
            os.mkdir(os.path.join(self.previous_root_path, "3_派工單_excel"))

        self.load_open_path = os.path.join(self.previous_root_path, "1_ERP產品_excel")

        self.load_file = os.path.join(self.load_open_path, listdir(self.load_open_path)[0])
        self.wb = load_workbook(self.load_file)
        self.sh = self.wb[self.wb.sheetnames[0]]

        self.db_file_1 = db_file_1
        self.db_file_2 = db_file_2

    def close(self):
        self.wb.close()

    def get_db_inform(self):

        db_1_2_config = []
        try:
            with open(self.db_file_1, 'r', encoding='utf-8') as fp_1:
                config_1 = load(fp_1)
            db_1_2_config.append(config_1['PATH'])
        except:
            with open(self.db_file_1, 'r', encoding='utf-8-sig') as fp_1:
                config_1 = load(fp_1)
            db_1_2_config.append(config_1['PATH'])

        with open(self.db_file_2, 'r') as fp_2:
            config_2 = load(fp_2)
        db_1_2_config.append(config_2['DBCONFIG']['HOST'])
        db_1_2_config.append(config_2['DBCONFIG']['USER'])
        db_1_2_config.append(config_2['DBCONFIG']['PASS'])
        db_1_2_config.append(config_2['DBCONFIG']['DATABASE'])

        return db_1_2_config

    def get_working_hour(self):
        db_setting = self.get_db_inform()

        try:
            cnx = connect(host=db_setting[1],
                          user=db_setting[2],
                          passwd=db_setting[3],
                          database=db_setting[4])
            if cnx.is_connected():
                db_info = cnx.get_server_info()
                cursor = cnx.cursor(named_tuple=True)
                print('db_version:', db_info)
        except Error as e:
            print("資料庫連接失敗：", e)
        cursor.execute("SELECT product_no,time FROM products")
        products = cursor.fetchall()
        cursor.close()
        cnx.close()
        return products

    def read_data(self):
        sheet = self.sh
        rows = list(sheet.rows)
        titles = []
        for t in rows[0]:
            title = t.value
            titles.append(title)
        cases = []
        for row in rows[1:]:
            case = []
            for r in row:
                case.append(r.value)
            cases.append(dict(zip(titles, case)))
        self.close()

        return cases

    def make_data(self):  # 準備在這邊寫讀取後的資料處理

        data = self.read_data()
        bom_path = self.get_db_inform()
        db = self.get_working_hour()

        total = []
        for row in data:
            number = str(row['(欄號)'])
            customer = str(row['客戶'])
            parent = str(row['產品編號'])
            name = str(row['品名規格'])
            amount = int(row['數量'])
            description = str(row['分錄備註'])
            customized_1 = str(row['自訂欄一'])
            customized_2 = str(row['自訂欄二'])
            single_wh = []  # 單件工時
            working_hour = []  # 工時
            process = []  # 製程
            dispatch_list = ''  # 對方填
            manu_order_num = ''  # 對方填

            tem_process = []

            # 拉出相對應的工時的部份
            for product in db:
                if str(product.product_no) == parent:  # 根據母件編號
                    single_wh.append(int(product.time))
                    working_hour.append(int(product.time * amount))

            if len(single_wh) == 0:  # 如果最後len是0就代表db裡面沒有相關資料，回傳NAN
                print('Cannot find matched data')
                single_wh.append(999999)
                working_hour.append(999999)

            # 找BOM製程的部份
            full_path = os.path.join(bom_path[0], parent[0], parent)

            try:
                xlrd_wb = open_workbook(full_path + '.xls')
            except FileNotFoundError as e:
                print(e)
                try:
                    xlrd_wb = open_workbook(full_path + '.xlsx')
                except:
                    xlrd_wb = False

            if xlrd_wb is not False:

                table = xlrd_wb.sheets()[0]

                prgm = set()
                lasr = set()
                bend = set()
                weld = set()

                for i in range(2, table.nrows):
                    prgm.add(str(table.row_values(i)[8]))

                for i in range(2, table.nrows):
                    lasr.add(str(table.row_values(i)[9]))

                for i in range(2, table.nrows):
                    bend.add(str(table.row_values(i)[10]))

                for i in range(2, table.nrows):
                    weld.add(str(table.row_values(i)[11]))

                if 'Y' in prgm:  # 將製程轉成MA,MB,MC,MD
                    tem_process.append('MA')
                if 'Y' in lasr:
                    tem_process.append('MB')
                if 'Y' in bend:
                    tem_process.append('MC')
                if 'Y' in weld:
                    tem_process.append('MD')

                process.append(" ".join(tem_process))
            else:
                process.append("NAN")

            total.append([number, customer, parent, name, amount, description, customized_1, customized_2,
                          single_wh[0], working_hour[0], process[0], dispatch_list, manu_order_num])

        total_sort = sorted(total, key=itemgetter(10, 8), reverse=True)
        return total_sort

    def write_data(self):
        data1 = self.make_data()
        rename = listdir(self.load_open_path)[0]

        workbook = Workbook()
        sheet = workbook[workbook.sheetnames[0]]

        sheet.cell(row=1, column=1, value='(欄號)').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=2, value='客戶').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=3, value='產品編號').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=4, value='品名規格').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=5, value='數量').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=6, value='分錄備註').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=7, value='自訂欄一').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=8, value='自訂欄二').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=9, value='單件工時').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=10, value='工時').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=11, value='製品製程').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=12, value='派工單').alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=13, value='製令批號').alignment = Alignment(horizontal='center')

        row_ = 2
        for data in data1:
            column_ = 1
            for item in data:
                cell = sheet.cell(row=row_, column=column_, value=item)
                cell.alignment = Alignment(horizontal='center')
                column_ += 1
            row_ += 1

        save_file = os.path.join(self.previous_root_path, "2_炸製令結果_excel", '{}.xlsx'.format(rename.split('.')[0]))
        workbook.save(save_file)
        workbook.close()

        pass


if __name__ == '__main__':
    excel = ReadExcel('config_bom.json', 'config_working_hour.json')  # 讀取檔案
    excel.write_data()   # 輸出新檔案