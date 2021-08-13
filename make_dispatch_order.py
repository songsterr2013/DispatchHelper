from os import listdir
import os.path
import sys
import xlwings as xw
from openpyxl import load_workbook
from datetime import datetime


class DispatchTable:

    def __init__(self, template_name):  # 傳入的參數為炸製令後excel

        self.root_path = ''

        if getattr(sys, 'frozen', False):  # 如果為exe
            self.root_path = os.path.dirname(sys.executable)
        elif __file__:
            self.root_path = os.path.dirname(__file__)

        self.previous_root_path = os.path.dirname(self.root_path)  # 取得檔案所在的前一路徑
        self.load_open_path = os.path.join(self.previous_root_path, "2_炸製令結果_excel")

        self.load_file = os.path.join(self.load_open_path, listdir(self.load_open_path)[0])
        self.wb = load_workbook(self.load_file)
        self.sh = self.wb[self.wb.sheetnames[0]]

        self.app = xw.App(visible=False, add_book=False)  # activate它

        self.template = self.app.books.open(template_name)  # 打開模板
        self.template_sheet = self.template.sheets[0]

        self.output = self.app.books.add()  # 創新excel 並貼上這邊
        self.output_sheet = self.output.sheets[0]

        self.written_row = 1  # 定義從第一行開始寫

    def save(self):
        rename = listdir(self.load_open_path)[0]
        save_file = os.path.join(self.previous_root_path, "3_派工單_excel", '{}.xlsx'.format(rename.split('.')[0]))
        self.output.save(save_file)

    def quit(self):
        self.template.close()
        self.output.close()
        self.app.quit()

    def write_header(self):  # 複製模板-頭
        self.template_sheet.range('A1:L6').copy(self.output_sheet.range('A{}'.format(self.written_row)))  # 貼過去
        self.written_row += 6  # header佔6行故+6

    def write_footer(self):  # 複製模板-尾
        self.template_sheet.range('A8:L29').copy(self.output_sheet.range('A{}'.format(self.written_row)))  # 貼過去
        self.written_row += 21  # footer佔21行故+21
        self.written_row += 2  # 跳2行繼續寫

    def read_data(self):  # 讀取炸製令後的excel

        sheet = self.sh
        rows = list(sheet.rows)

        titles = []
        for t in rows[0]:
            title = t.value
            titles.append(title)
        cases = []
        for row in rows[1:]:
            case = []
            for r in row:
                case.append(r.value)
            cases.append(dict(zip(titles, case)))
        self.wb.close()

        return cases  # return出一個list

    def rearrange_template_2(self):  # 讀取新的excel, 處理合併cell為空值的情況且排序

        pre_data = self.read_data()
        data = []

        for i in pre_data:  # 只拿有用的DATA
            if i['(欄號)'] is not None:
                data.append(i)

        for dic in data:
            if dic['派工單'] is not None:
                dispatch_id = dic['派工單']
            else:
                dic['派工單'] = dispatch_id

            if dic['製令批號'] is not None:
                batch_number = dic['製令批號']
            else:
                dic['製令批號'] = batch_number

            if dic['自訂欄一'] is None:
                dic['自訂欄一'] = ''

            if dic['自訂欄二'] is None:
                dic['自訂欄二'] = ''

        sorted_data = sorted(data, key=lambda x: x['派工單'])  # 排序

        return sorted_data

    def run(self):

        data = self.rearrange_template_2()
        for amount in range(1, int(data[-1]['派工單'].split('-')[0].split('#')[1]) + 1):  # 根據派工單的份數製造派工單

            # 1 接表頭
            self.write_header()

            # 2 填行
            first_time = 0  # 讓它只跑第一圈，資料拿一次就好
            parent_amounts = 0  # 母件數量
            etm_time = 0  # 預計工時
            data_row = 0  # 這一張派工單有幾行data

            for row in data:
                # 由於派工單那一欄的CELL裡面可能會填到(白鐵、鍍鋅、染黑)，所以需要這樣的方法取得那個份數
                if str(row['派工單']).endswith(')'):
                    number = row['派工單'].split('-')[1].split('(')[0]
                else:
                    number = row['派工單'].split('-')[1]

                if int(number) == amount:
                    if first_time == 0:
                        if str(row['派工單']).endswith(')'):
                            if str(row['派工單'].split('(')[1][0:3]) == 'SUS':
                                title_1 = str(row['派工單']).split('(')[0] + ' 焊 接 課 個 人 工 作 日 報 表(白鐵)'
                            elif str(row['派工單'].split('(')[1][0:2]) == 'BK':
                                title_1 = str(row['派工單']).split('(')[0] + ' 焊 接 課 個 人 工 作 日 報 表(染黑)'
                            elif str(row['派工單'].split('(')[1][0:2]) == 'ZN':
                                title_1 = str(row['派工單']).split('(')[0] + ' 焊 接 課 個 人 工 作 日 報 表(鍍鋅)'
                        else:
                            title_1 = str(row['派工單']) + ' 焊 接 課 個 人 工 作 日 報 表'
                        title_2 = '製令單號:' + str(row['製令批號']) + '/' + str(row['自訂欄一']) + '/' + str(row['自訂欄二'])
                        date = '登打: ' + str(datetime.now().strftime('%m/%d'))
                        client = row['客戶']

                        first_time += 1  # 讓它只跑第一圈

                    self.output_sheet.range('A{}'.format(self.written_row)).value = row['產品編號']
                    self.output_sheet.range('A{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('A{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('A{}'.format(self.written_row)).api.HorizontalAlignment = -4108

                    self.output_sheet.range('B{}'.format(self.written_row)).value = row['品名規格']
                    self.output_sheet.range('B{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('B{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('B{}'.format(self.written_row)).api.HorizontalAlignment = -4108

                    self.output_sheet.range('C{}'.format(self.written_row)).value = row['數量']
                    self.output_sheet.range('C{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('C{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('C{}'.format(self.written_row)).api.HorizontalAlignment = -4108

                    self.output_sheet.range('D{}'.format(self.written_row)).value = row['工時']
                    self.output_sheet.range('D{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('D{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('D{}'.format(self.written_row)).api.HorizontalAlignment = -4108

                    self.output_sheet.range('E{}'.format(self.written_row)).api.Borders(10).LineStyle = 1  # blank
                    self.output_sheet.range('E{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('G{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('G{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('I{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('I{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('L{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('L{}'.format(self.written_row)).api.Borders(10).Weight = 3

                    self.output_sheet.range('J{}'.format(self.written_row)).value = '□'
                    self.output_sheet.range('J{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('J{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('J{}'.format(self.written_row)).api.HorizontalAlignment = -4108

                    self.output_sheet.range('K{}'.format(self.written_row)).value = '□'
                    self.output_sheet.range('K{}'.format(self.written_row)).api.Borders(10).LineStyle = 1
                    self.output_sheet.range('K{}'.format(self.written_row)).api.Borders(10).Weight = 3
                    self.output_sheet.range('K{}'.format(self.written_row)).api.HorizontalAlignment = -4108

                    parent_amounts += int(row['數量'])  # 母件數量
                    etm_time += int(row['工時'])  # 預計工時

                    self.written_row += 1  # 沿著現在的東西一行一行地寫
                    data_row += 1

            # 3 接表腳
            self.write_footer()

            # 4 填表
            self.output_sheet.range('A{}'.format(self.written_row - data_row - 2 - 27)).value = title_1
            self.output_sheet.range('A{}'.format(self.written_row - data_row - 2 - 26)).value = title_2
            self.output_sheet.range('J{}'.format(self.written_row - data_row - 2 - 24)).value = date
            self.output_sheet.range('L{}'.format(self.written_row - data_row - 2 - 23)).value = client

            self.output_sheet.range('A{}'.format(self.written_row - 2 - 20)).value = parent_amounts
            self.output_sheet.range('C{}'.format(self.written_row - 2 - 20)).value = etm_time
            self.output_sheet.range('A{}'.format(self.written_row - 2 - 10)).value = title_1
            self.output_sheet.range('A{}'.format(self.written_row - 2 - 9)).value = title_2
            self.output_sheet.range('B{}'.format(self.written_row - 2 - 6)).value = parent_amounts

            if title_1[-1] == ')':
                self.output_sheet.range('A{}'.format(self.written_row - data_row - 2 - 27)).api.Font.Color = 0x0000ff
                self.output_sheet.range('A{}'.format(self.written_row - 2 - 10)).api.Font.Color = 0x0000ff


if __name__ == '__main__':
    dispatch_table = DispatchTable('template.xlsx')
    dispatch_table.run()
    dispatch_table.save()
    dispatch_table.quit()
